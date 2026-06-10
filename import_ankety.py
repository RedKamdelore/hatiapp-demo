#!/usr/bin/env python3
"""
Импорт волонтёров из Excel-анкет.
Обновляет arrival_date, departure_date, full_name, is_active.
"""
import sys
sys.path.insert(0, r'C:\Users\Administrator\Desktop\САМОПАЛ\Hatiapp_cowork_OPENCODE')

import openpyxl
from datetime import date, datetime, timedelta
from database import SessionLocal
import models
import re

# Даты расписания в БД
SCHEDULE_START = date(2026, 7, 9)
SCHEDULE_END = date(2026, 7, 13)

def parse_date_from_header(header):
    """Извлекает дату из заголовка колонки."""
    if not header or not isinstance(header, str):
        return None
    
    # Ищем паттерн "число месяца"
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    
    # Паттерн: "4 июля" или "10 июля"
    match = re.search(r'(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)', header.lower())
    if match:
        day = int(match.group(1))
        month = months[match.group(2)]
        return date(2026, month, day)
    
    return None

def get_first_last_yes(row_data, date_columns):
    """Возвращает (arrival_date, departure_date) на основе ответов 'Да'."""
    yes_dates = []
    
    for col_name, cell_value in row_data.items():
        if col_name in date_columns:
            d = date_columns[col_name]
            if cell_value and str(cell_value).strip().lower() == 'да':
                yes_dates.append(d)
    
    if not yes_dates:
        return None, None
    
    yes_dates.sort()
    arrival = yes_dates[0]
    departure = yes_dates[-1] + timedelta(days=1)
    
    return arrival, departure

def process_excel():
    db = SessionLocal()
    
    # Загружаем Excel
    wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\Анкеты Хати 2026 (1).xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    # Карта колонка -> дата
    date_columns = {}
    for i, h in enumerate(headers):
        d = parse_date_from_header(h)
        if d:
            date_columns[h] = d
    
    print(f"Найдено колонок с датами: {len(date_columns)}")
    for h, d in sorted(date_columns.items(), key=lambda x: x[1]):
        print(f"  {d}: {h}")
    
    updated = 0
    created = 0
    blocked_no_dates = []
    blocked_out_of_schedule = []
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
        
        username = row_data.get('Позывной')
        if not username:
            continue
        
        username = str(username).strip()
        
        # Имя с большой буквы
        full_name = username.capitalize()
        
        # Находим даты
        arrival, departure = get_first_last_yes(row_data, date_columns)
        
        if not arrival or not departure:
            # Нет дат — блокируем
            user = db.query(models.User).filter(models.User.username == username).first()
            if user:
                user.is_active = False
                blocked_no_dates.append(username)
            continue
        
        # Проверяем попадание на расписание
        # Человек должен быть в лагере хотя бы часть расписания
        # arrival < SCHEDULE_END и departure > SCHEDULE_START
        if arrival >= SCHEDULE_END or departure <= SCHEDULE_START:
            user = db.query(models.User).filter(models.User.username == username).first()
            if user:
                user.is_active = False
                reason = f"не попадает на расписание (заезд {arrival}, отъезд {departure})"
                blocked_out_of_schedule.append(f"{username}: {reason}")
            continue
        
        # Ищем или создаём пользователя
        user = db.query(models.User).filter(models.User.username == username).first()
        
        if user:
            # Обновляем
            user.full_name = full_name
            user.arrival_date = arrival
            user.departure_date = departure
            user.is_active = True
            updated += 1
        else:
            # Создаём нового
            from services.auth import hash_password
            import secrets
            temp_password = secrets.token_urlsafe(8)
            
            user = models.User(
                username=username.lower(),
                full_name=full_name,
                password_hash=hash_password(temp_password),
                role='volunteer',
                is_active=True,
                arrival_date=arrival,
                departure_date=departure,
            )
            db.add(user)
            created += 1
            print(f"Создан новый пользователь: {username} (пароль: {temp_password})")
    
    db.commit()
    
    print("\n" + "="*60)
    print("РЕЗУЛЬТАТЫ:")
    print(f"  Обновлено: {updated}")
    print(f"  Создано: {created}")
    print(f"  Заблокировано (нет дат): {len(blocked_no_dates)}")
    if blocked_no_dates:
        print(f"    {', '.join(blocked_no_dates)}")
    print(f"  Заблокировано (вне расписания): {len(blocked_out_of_schedule)}")
    for item in blocked_out_of_schedule:
        print(f"    • {item}")
    print("="*60)
    
    db.close()

if __name__ == "__main__":
    process_excel()
