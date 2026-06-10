import io
import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from datetime import date
import models
from services.booking import get_slot_stats


def generate_qr(data: str) -> bytes:
    """Генерирует QR-код и возвращает PNG байты."""
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def export_schedule_excel(target_date: date, db: Session) -> bytes:
    """
    Экспортирует расписание на конкретный день в Excel.
    Возвращает байты файла.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = str(target_date)

    # --- Стили ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # индиго
    center = Alignment(horizontal="center", vertical="center")

    # --- Заголовки ---
    headers = ["Направление", "Время", "Мест всего", "Занято", "Свободно", "Волонтёры"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # --- Данные ---
    directions = db.query(models.Direction).order_by(models.Direction.name).all()
    row = 2
    for direction in directions:
        slots = db.query(models.Slot).filter_by(
            direction_id=direction.id, date=target_date
        ).order_by(models.Slot.time).all()

        for slot in slots:
            if slot.capacity == 0:
                continue
            stats = get_slot_stats(slot, db)
            volunteers = ", ".join(
                b.user.full_name or b.user.username
                for b in slot.bookings
            )
            ws.append([
                direction.name,
                slot.time.strftime("%H:%M"),
                slot.capacity,
                stats["booked"],
                stats["free"],
                volunteers,
            ])
            row += 1

    # --- Ширина колонок ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
