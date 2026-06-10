#!/usr/bin/env python3
"""
Обновление существующих пользователей датами из Excel.
Для 11 дубликатов (leader/admin/lotos) берём даты заезда/отъезда из анкет.
"""
import sys
sys.path.insert(0, r'C:\Users\Administrator\Desktop\САМОПАЛ\Hatiapp_cowork_OPENCODE')

import openpyxl
from datetime import date, timedelta
from database import SessionLocal
import models
import re

SCHEDULE_START = date(2026, 7, 9)
SCHEDULE_END = date(2026, 7, 13)

def parse_date_from_header(header):
    if not header or not isinstance(header, str):
        return None
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    match = re.search(r'(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)', header.lower())
    if match:
        day = int(match.group(1))
        month = months[match.group(2)]
        return date(2026, month, day)
    return None

def get_dates(row, date_columns, headers):
    yes_dates = []
    for col_name, d in date_columns.items():
        col_idx = headers.index(col_name)
        if col_idx < len(row):
            val = row[col_idx]
            if val and str(val).strip().lower() == 'да':
                yes_dates.append(d)
    if not yes_dates:
        return None, None
    yes_dates.sort()
    return yes_dates[0], yes_dates[-1] + timedelta(days=1)

def process():
    db = SessionLocal()
    wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\Анкеты Хати 2026 (1).xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    date_columns = {}
    for h in headers:
        d = parse_date_from_header(h)
        if d:
            date_columns[h] = d
    
    updated = 0
    not_found = []
    skipped = []
    
    from services.auth import hash_password
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
            
        pozyvnoy = row[0]
        phone = row[4]
        tg = row[5]
        
        if not tg or not pozyvnoy:
            continue
        
        username = str(tg).strip()
        full_name = str(pozyvnoy).strip()
        password = str(phone).strip() if phone else 'changeme'
        
        # Ищем пользователя в БД
        user = db.query(models.User).filter(models.User.username == username).first()
        
        if not user:
            not_found.append(f"{username} ({full_name}): не найден в БД")
            continue
        
        # Если уже волонтёр — пропускаем (уже создан с датами)
        if user.role == 'volunteer':
            skipped.append(f"{username}: уже волонтёр, пропускаем")
            continue
        
        # Получаем даты
        arrival, departure = get_dates(row, date_columns, headers)
        
        if not arrival or not departure:
            skipped.append(f"{username}: нет дат в Excel")
            continue
        
        # Обновляем существующего пользователя
        user.full_name = full_name
        user.password_hash = hash_password(password)
        user.arrival_date = arrival
        user.departure_date = departure
        user.is_active = True
        updated += 1
        print(f"Обновлён: {username} ({user.role}) — arrival={arrival}, departure={departure}")
    
    db.commit()
    
    print("\n" + "="*60)
    print(f"Обновлено: {updated}")
    print(f"Пропущено (уже волонтёры): {len(skipped)}")
    print(f"Не найдены: {len(not_found)}")
    if not_found:
        for n in not_found:
            print(f"  • {n}")
    print("="*60)
    
    # Проверим итог
    total_volunteers = db.query(models.User).filter(models.User.role == 'volunteer').count()
    total_with_dates = db.query(models.User).filter(models.User.arrival_date != None).count()
    print(f"\nИтого волонтёров: {total_volunteers}")
    print(f"Пользователей с датами (все роли): {total_with_dates}")
    
    db.close()

if __name__ == "__main__":
    process()
