#!/usr/bin/env python3
"""
Импорт волонтёров из Excel-анкет (исправленный v2).
username = ТГ (колонка 5), full_name = Позывной (колонка 0), password = Телефон (колонка 4).
"""
import sys
sys.path.insert(0, r'C:\Users\Administrator\Desktop\САМОПАЛ\Hatiapp_cowork_OPENCODE')

import openpyxl
from datetime import date, timedelta
from database import SessionLocal
import models
import re

SCHEDULE_START = date(2026, 7, 9)
SCHEDULE_END = date(2026, 7, 13)

def parse_date_from_header(header):
    if not header or not isinstance(header, str):
        return None
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    match = re.search(r'(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)', header.lower())
    if match:
        day = int(match.group(1))
        month = months[match.group(2)]
        return date(2026, month, day)
    return None

def get_dates(row_data, date_columns):
    yes_dates = []
    for col_name, d in date_columns.items():
        val = row_data.get(col_name)
        if val and str(val).strip().lower() == 'да':
            yes_dates.append(d)
    if not yes_dates:
        return None, None
    yes_dates.sort()
    return yes_dates[0], yes_dates[-1] + timedelta(days=1)

def process():
    db = SessionLocal()
    wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\Анкеты Хати 2026 (1).xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    # Карта колонок с датами
    date_columns = {}
    for h in headers:
        d = parse_date_from_header(h)
        if d:
            date_columns[h] = d
    
    print(f"Найдено колонок с датами: {len(date_columns)}")
    
    created = 0
    updated = 0
    skipped = 0
    blocked = []
    password_map = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
        
        # Используем индексы колонок
        pozyvnoy = row[0] if len(row) > 0 else None
        phone = row[4] if len(row) > 4 else None
        tg = row[5] if len(row) > 5 else None
        
        if not tg or not pozyvnoy:
            skipped += 1
            continue
        
        username = str(tg).strip().lower()
        full_name = str(pozyvnoy).strip()
        password = str(phone).strip() if phone else 'changeme'
        
        arrival, departure = get_dates(row_data, date_columns)
        
        if not arrival or not departure:
            blocked.append(f"{username} ({full_name}): нет дат")
            continue
        
        if arrival >= SCHEDULE_END or departure <= SCHEDULE_START:
            blocked.append(f"{username} ({full_name}): вне расписания")
            continue
        
        # Ищем существующего
        user = db.query(models.User).filter(models.User.username == username).first()
        
        from services.auth import hash_password
        
        if user:
            user.full_name = full_name
            user.password_hash = hash_password(password)
            user.arrival_date = arrival
            user.departure_date = departure
            user.is_active = True
            updated += 1
        else:
            user = models.User(
                username=username,
                full_name=full_name,
                password_hash=hash_password(password),
                role='volunteer',
                is_active=True,
                arrival_date=arrival,
                departure_date=departure,
            )
            db.add(user)
            created += 1
        
        password_map.append((username, password, full_name, arrival, departure))
    
    db.commit()
    
    # Сохраняем пароли
    with open('volunteer_passwords.txt', 'w', encoding='utf-8') as f:
        f.write('username|password|full_name|arrival|departure\n')
        for item in password_map:
            f.write('|'.join(str(x) for x in item) + '\n')
    
    print("\n" + "="*60)
    print(f"Создано: {created}")
    print(f"Обновлено: {updated}")
    print(f"Пропущено: {skipped}")
    print(f"Заблокировано: {len(blocked)}")
    for b in blocked[:10]:
        print(f"  • {b}")
    if len(blocked) > 10:
        print(f"  ... и ещё {len(blocked) - 10}")
    print("="*60)
    
    db.close()

if __name__ == "__main__":
    process()
