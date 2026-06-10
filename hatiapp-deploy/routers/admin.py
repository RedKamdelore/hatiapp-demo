from fastapi import APIRouter, Request, Depends, Query, Form
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from datetime import datetime, date
import math

from database import get_db
from services.auth import require_role
from services.export import export_schedule_excel, generate_qr
from config import ROLE_ADMIN, ROLE_LEADER, ROLE_VOLUNTEER, ROLE_LOTOS, ROLE_PERMANENT, BOOKINGS_PER_DAY
import models

# ── Палитры цветовых тем ────────────────────────────────────────────────────
THEMES: dict[str, dict] = {
    "indigo": {
        "label": "Индиго",
        "50":  "#eef2ff", "100": "#e0e7ff", "200": "#c7d2fe",
        "300": "#a5b4fc", "400": "#818cf8", "500": "#6366f1",
        "600": "#4f46e5", "700": "#4338ca", "800": "#3730a3", "900": "#312e81",
    },
    "violet": {
        "label": "Фиолетовый",
        "50":  "#f5f3ff", "100": "#ede9fe", "200": "#ddd6fe",
        "300": "#c4b5fd", "400": "#a78bfa", "500": "#8b5cf6",
        "600": "#7c3aed", "700": "#6d28d9", "800": "#5b21b6", "900": "#4c1d95",
    },
    "blue": {
        "label": "Синий",
        "50":  "#eff6ff", "100": "#dbeafe", "200": "#bfdbfe",
        "300": "#93c5fd", "400": "#60a5fa", "500": "#3b82f6",
        "600": "#2563eb", "700": "#1d4ed8", "800": "#1e40af", "900": "#1e3a8a",
    },
    "teal": {
        "label": "Бирюзовый",
        "50":  "#f0fdfa", "100": "#ccfbf1", "200": "#99f6e4",
        "300": "#5eead4", "400": "#2dd4bf", "500": "#14b8a6",
        "600": "#0d9488", "700": "#0f766e", "800": "#115e59", "900": "#134e4a",
    },
    "emerald": {
        "label": "Изумрудный",
        "50":  "#ecfdf5", "100": "#d1fae5", "200": "#a7f3d0",
        "300": "#6ee7b7", "400": "#34d399", "500": "#10b981",
        "600": "#059669", "700": "#047857", "800": "#065f46", "900": "#064e3b",
    },
    "rose": {
        "label": "Розовый",
        "50":  "#fff1f2", "100": "#ffe4e6", "200": "#fecdd3",
        "300": "#fda4af", "400": "#fb7185", "500": "#f43f5e",
        "600": "#e11d48", "700": "#be123c", "800": "#9f1239", "900": "#881337",
    },
    "amber": {
        "label": "Янтарный",
        "50":  "#fffbeb", "100": "#fef3c7", "200": "#fde68a",
        "300": "#fcd34d", "400": "#fbbf24", "500": "#f59e0b",
        "600": "#d97706", "700": "#b45309", "800": "#92400e", "900": "#78350f",
    },
}

def _get_setting(db: Session, key: str, default: str = "") -> str:
    row = db.query(models.AppSetting).filter_by(key=key).first()
    return row.value if row else default

def _set_setting(db: Session, key: str, value: str) -> None:
    row = db.query(models.AppSetting).filter_by(key=key).first()
    if row:
        row.value = value
    else:
        db.add(models.AppSetting(key=key, value=value))
    db.commit()

router = APIRouter()
templates = Jinja2Templates(directory="templates")


@router.get("/admin", response_class=HTMLResponse)
def admin_panel(
    request: Request,
    sort: str = Query(default="role_asc"),
    db: Session = Depends(get_db),
):
    user = require_role(request, db, ROLE_ADMIN)

    directions = db.query(models.Direction).options(
        joinedload(models.Direction.leaders).joinedload(models.DirectionLeader.user)
    ).order_by(models.Direction.name).all()

    # Сортировка пользователей
    query = db.query(models.User)
    if sort == "name_asc":
        query = query.order_by(models.User.full_name.asc().nullslast(), models.User.username.asc())
    elif sort == "name_desc":
        query = query.order_by(models.User.full_name.desc().nullsfirst(), models.User.username.desc())
    elif sort == "username_asc":
        query = query.order_by(models.User.username.asc())
    elif sort == "username_desc":
        query = query.order_by(models.User.username.desc())
    elif sort == "role_asc":
        query = query.order_by(models.User.role.asc(), models.User.username.asc())
    elif sort == "role_desc":
        query = query.order_by(models.User.role.desc(), models.User.username.asc())
    elif sort == "status_asc":
        query = query.order_by(models.User.is_active.desc(), models.User.username.asc())
    elif sort == "status_desc":
        query = query.order_by(models.User.is_active.asc(), models.User.username.asc())
    else:
        query = query.order_by(models.User.role.asc(), models.User.username.asc())
    users = query.all()

    # Все слоты одним запросом
    all_slots = db.query(models.Slot).all()
    slots_by_dir: dict = {}
    for s in all_slots:
        slots_by_dir.setdefault(s.direction_id, []).append(s)

    # Количество броней по слотам одним запросом
    booking_counts = dict(
        db.query(models.Booking.slot_id, func.count(models.Booking.id))
        .group_by(models.Booking.slot_id).all()
    )

    # Количество явок по слотам одним запросом
    attended_by_slot = dict(
        db.query(models.Booking.slot_id, func.count(models.Attendance.id))
        .join(models.Attendance, models.Attendance.booking_id == models.Booking.id)
        .filter(models.Attendance.present == True)
        .group_by(models.Booking.slot_id).all()
    )

    bookings_count = sum(booking_counts.values())

    # Счётчики для статистики
    total_capacity = sum(s.capacity for s in all_slots)
    total_booked   = sum(booking_counts.get(s.id, 0) for s in all_slots)
    available_slots = total_capacity - total_booked
    total_shifts    = len(all_slots)
    
    # Человеко-смены: сколько должен записаться каждый волонтёр (бессменные не считаются)
    volunteer_count = len([u for u in users if u.role == ROLE_VOLUNTEER])
    permanent_count = len([u for u in users if u.role == ROLE_PERMANENT])
    schedule_days_count = len(set(s.date for s in all_slots))
    total_human_shifts = volunteer_count * BOOKINGS_PER_DAY * schedule_days_count

    dir_stats = []
    for d in directions:
        slots = slots_by_dir.get(d.id, [])
        total    = sum(s.capacity for s in slots)
        booked   = sum(booking_counts.get(s.id, 0) for s in slots)
        attended = sum(attended_by_slot.get(s.id, 0) for s in slots)
        dir_stats.append({
            "direction": d,
            "total": total,
            "booked": booked,
            "attended": attended,
        })

    current_theme = _get_setting(db, "theme", "indigo")

    # Заблокированные дни
    blocked_days = db.query(models.BlockedDay).order_by(models.BlockedDay.date.desc()).all()

    # Данные для графика посещаемости по дням + нехватка людей
    attendance_chart = []
    day_coverage_stats = []
    all_dates = sorted(set(s.date for s in all_slots))
    for d in all_dates:
        day_slots = [s for s in all_slots if s.date == d]
        day_capacity = sum(s.capacity for s in day_slots)
        day_booked = sum(booking_counts.get(s.id, 0) for s in day_slots)
        day_attended = sum(attended_by_slot.get(s.id, 0) for s in day_slots)
        attendance_chart.append({
            "date": d.isoformat(),
            "date_fmt": d.strftime("%d.%m"),
            "booked": day_booked,
            "attended": day_attended,
        })
        
        # Сколько волонтёров присутствует в этот день (бессменные работают отдельно)
        present_count = 0
        for u in users:
            if u.role == ROLE_VOLUNTEER and u.is_active:
                if u.arrival_date and u.departure_date:
                    if u.arrival_date < d < u.departure_date:
                        present_count += 1
                else:
                    present_count += 1
        
        # Формула: (доступные люди × 2 смены) — хватает ли перекрыть все слоты
        needed_people = math.ceil(day_capacity / BOOKINGS_PER_DAY) if day_capacity > 0 else 0
        shortage_people = max(0, needed_people - present_count)
        coverage_ok = shortage_people == 0
        
        day_coverage_stats.append({
            "date": d,
            "date_fmt": d.strftime("%d.%m"),
            "capacity": day_capacity,
            "booked": day_booked,
            "attended": day_attended,
            "present": present_count,
            "needed_people": needed_people,
            "shortage_people": shortage_people,
            "coverage_ok": coverage_ok,
        })

    # Все даты из расписания для выпадающего списка
    schedule_dates = sorted(set(
        row[0] for row in db.query(models.Slot.date).distinct().all()
    ))

    return templates.TemplateResponse("admin.html", {
        "request": request,
        "user": user,
        "users": users,
        "dir_stats": dir_stats,
        "bookings_count": bookings_count,
        "available_slots": available_slots,
        "total_capacity": total_capacity,
        "total_human_shifts": total_human_shifts,
        "today": date.today().isoformat(),
        "current_theme": current_theme,
        "blocked_days": blocked_days,
        "schedule_dates": schedule_dates,
        "attendance_chart": attendance_chart,
        "day_coverage_stats": day_coverage_stats,
        "current_sort": sort,
    })


@router.get("/admin/export")
def export_excel(
    request: Request,
    day: str = Query(default=None),
    db: Session = Depends(get_db),
):
    require_role(request, db, ROLE_ADMIN)
    target = datetime.strptime(day, "%Y-%m-%d").date() if day else date.today()
    data = export_schedule_excel(target, db)
    filename = f"schedule_{target}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/admin/print", response_class=HTMLResponse)
def print_schedule(
    request: Request,
    day: str = Query(default=None),
    db: Session = Depends(get_db),
):
    """Страница для печати расписания на конкретный день."""
    require_role(request, db, ROLE_ADMIN)
    target = datetime.strptime(day, "%Y-%m-%d").date() if day else date.today()

    # Русские названия дней недели
    WEEKDAYS_RU = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
    weekday = WEEKDAYS_RU[target.weekday()]

    directions = db.query(models.Direction).order_by(models.Direction.name).all()
    booking_counts = dict(
        db.query(models.Booking.slot_id, func.count(models.Booking.id))
        .group_by(models.Booking.slot_id).all()
    )

    directions_data = []
    total_booked = 0
    total_free = 0
    total_slots = 0
    dir_count = 0

    for direction in directions:
        slots = db.query(models.Slot).filter_by(
            direction_id=direction.id, date=target
        ).filter(models.Slot.capacity > 0).order_by(models.Slot.time).all()
        if not slots:
            continue
        dir_count += 1
        dir_rows = []
        for idx, slot in enumerate(slots):
            booked = booking_counts.get(slot.id, 0)
            free = slot.capacity - booked
            pct = int(booked / slot.capacity * 100) if slot.capacity else 0
            names = ", ".join(
                b.user.full_name or b.user.username
                for b in slot.bookings
            )
            dir_rows.append({
                "time":        slot.time.strftime("%H:%M"),
                "capacity":    slot.capacity,
                "booked":      booked,
                "free":        free,
                "pct":         pct,
                "names":       names,
            })
            total_booked += booked
            total_free   += free
            total_slots  += 1
        directions_data.append({
            "name": direction.name,
            "rows": dir_rows,
            "row_count": len(dir_rows),
        })

    # Распределяем направления по 3 столбикам примерно поровну
    columns = [[], [], []]
    col_sizes = [0, 0, 0]
    for d in directions_data:
        # Выбираем столбик с минимальным количеством строк
        min_col = col_sizes.index(min(col_sizes))
        columns[min_col].append(d)
        col_sizes[min_col] += d["row_count"]

    # Все даты из расписания для выпадающего списка
    schedule_dates = sorted(set(
        row[0] for row in db.query(models.Slot.date).distinct().all()
    ))

    return templates.TemplateResponse("print_schedule.html", {
        "request":     request,
        "target_date": target,
        "weekday":     weekday,
        "columns":     columns,
        "total_slots": total_slots,
        "total_booked": total_booked,
        "total_free":  total_free,
        "dir_count":   dir_count,
        "schedule_dates": schedule_dates,
    })


@router.get("/admin/qr/{booking_id}")
def booking_qr(booking_id: int, request: Request, db: Session = Depends(get_db)):
    from services.auth import get_current_user
    user = get_current_user(request, db)
    booking = db.query(models.Booking).filter_by(id=booking_id).first()
    if not booking:
        return Response(status_code=404)
    # Волонтёр может смотреть только свои QR, админ и руководитель — любые
    if user.role == ROLE_VOLUNTEER and booking.user_id != user.id:
        return Response(status_code=403)
    data = f"booking:{booking.id}:user:{booking.user_id}:slot:{booking.slot_id}"
    png = generate_qr(data)
    return Response(content=png, media_type="image/png")


@router.get("/api/qr/{booking_id}")
def api_qr(booking_id: int, request: Request, db: Session = Depends(get_db)):
    """Возвращает QR-код как base64 JSON для модального окна."""
    from services.auth import get_current_user
    import base64
    user = get_current_user(request, db)
    booking = db.query(models.Booking).filter_by(id=booking_id).first()
    if not booking:
        return JSONResponse({"error": "not found"}, status_code=404)
    if user.role == ROLE_VOLUNTEER and booking.user_id != user.id:
        return JSONResponse({"error": "forbidden"}, status_code=403)
    data = f"booking:{booking.id}:user:{booking.user_id}:slot:{booking.slot_id}"
    png = generate_qr(data)
    b64 = base64.b64encode(png).decode()
    return JSONResponse({
        "qr_b64": b64,
        "booking_id": booking.id,
        "direction": booking.slot.direction.name,
        "date": booking.slot.date.strftime("%d.%m.%Y"),
        "time": booking.slot.time.strftime("%H:%M"),
    })


@router.post("/admin/user/{user_id}/toggle")
def toggle_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    u = db.query(models.User).filter_by(id=user_id).first()
    if u:
        u.is_active = not u.is_active
        db.commit()
    return RedirectResponse("/admin", status_code=302)


@router.post("/admin/user/{user_id}/delete")
def delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    u = db.query(models.User).filter_by(id=user_id).first()
    if u and u.role != ROLE_ADMIN:
        db.delete(u)
        db.commit()
    return RedirectResponse("/admin", status_code=302)


@router.post("/admin/users/mass-action")
def mass_action_users(
    request: Request,
    db: Session = Depends(get_db),
    action: str = Form(...),
    user_ids: list[int] = Form(default=[]),
):
    require_role(request, db, ROLE_ADMIN)
    if not user_ids:
        return RedirectResponse("/admin?toast=Никто+не+выбран&toast_type=error", status_code=302)

    users = db.query(models.User).filter(models.User.id.in_(user_ids)).all()
    count = 0
    for u in users:
        if action == "activate":
            u.is_active = True
            count += 1
        elif action == "deactivate":
            if u.role != ROLE_ADMIN:
                u.is_active = False
                count += 1
        elif action == "delete":
            if u.role != ROLE_ADMIN:
                db.delete(u)
                count += 1
    db.commit()

    toast_msg = f"Обработано+{count}+пользователей"
    return RedirectResponse(f"/admin?toast={toast_msg}&toast_type=success", status_code=302)


@router.get("/admin/user/{user_id}/edit", response_class=HTMLResponse)
def edit_user_page(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    u = db.query(models.User).filter_by(id=user_id).first()
    if not u:
        return RedirectResponse("/admin", status_code=302)
    return templates.TemplateResponse("edit_user.html", {
        "request": request,
        "u": u,
        "error": None,
        "success": False,
    })


@router.post("/admin/user/{user_id}/edit", response_class=HTMLResponse)
def edit_user(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    username:  str = Form(...),
    full_name: str = Form(""),
    role:      str = Form(...),
    password:  str = Form(""),
    arrival_date:   str = Form(""),
    departure_date: str = Form(""),
):
    require_role(request, db, ROLE_ADMIN)
    u = db.query(models.User).filter_by(id=user_id).first()
    if not u:
        return RedirectResponse("/admin", status_code=302)

    # Проверяем уникальность логина
    conflict = db.query(models.User).filter(
        models.User.username == username,
        models.User.id != user_id,
    ).first()
    if conflict:
        return templates.TemplateResponse("edit_user.html", {
            "request": request,
            "u": u,
            "error": f"Логин «{username}» уже занят",
            "success": False,
        })

    u.username  = username.strip()
    u.full_name = full_name.strip() or None
    u.role      = role if role in (ROLE_ADMIN, ROLE_LEADER, ROLE_VOLUNTEER, ROLE_LOTOS, ROLE_PERMANENT) else ROLE_VOLUNTEER
    if password.strip():
        from services.auth import hash_password
        u.password_hash = hash_password(password.strip())

    # Парсим даты заезда/отъезда
    if arrival_date.strip():
        try:
            u.arrival_date = datetime.strptime(arrival_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            pass
    else:
        u.arrival_date = None

    if departure_date.strip():
        try:
            u.departure_date = datetime.strptime(departure_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            pass
    else:
        u.departure_date = None

    db.commit()
    return templates.TemplateResponse("edit_user.html", {
        "request": request,
        "u": u,
        "error": None,
        "success": True,
    })


@router.post("/admin/direction/{direction_id}/leaders/{user_id}")
def add_leader(
    direction_id: int,
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Добавить руководителя к направлению."""
    require_role(request, db, ROLE_ADMIN)
    d = db.query(models.Direction).filter_by(id=direction_id).first()
    u = db.query(models.User).filter_by(id=user_id, role=ROLE_LEADER).first()
    if d and u:
        # Проверяем, что связи ещё нет
        exists = db.query(models.DirectionLeader).filter_by(
            direction_id=direction_id, user_id=user_id
        ).first()
        if not exists:
            db.add(models.DirectionLeader(direction_id=direction_id, user_id=user_id))
            db.commit()
    return RedirectResponse("/admin", status_code=302)


@router.post("/admin/direction/{direction_id}/leaders/{user_id}/remove")
def remove_leader(
    direction_id: int,
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Убрать руководителя с направления."""
    require_role(request, db, ROLE_ADMIN)
    link = db.query(models.DirectionLeader).filter_by(
        direction_id=direction_id, user_id=user_id
    ).first()
    if link:
        db.delete(link)
        db.commit()
    return RedirectResponse("/admin", status_code=302)


# ---- Импорт пользователей из Excel ----
from fastapi import UploadFile, File
from services.import_users import import_users_from_excel


@router.get("/admin/import", response_class=HTMLResponse)
def import_page(request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    return templates.TemplateResponse("import_users.html", {"request": request, "result": None})


@router.post("/admin/import", response_class=HTMLResponse)
async def import_users(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    require_role(request, db, ROLE_ADMIN)
    contents = await file.read()
    result = import_users_from_excel(contents, db)
    return templates.TemplateResponse("import_users.html", {
        "request": request,
        "result": result,
    })


# ---- Импорт расписания из Excel ----
from seed.schedule import seed_schedule, _load_from_xlsx, SCHEDULE_XLSX
import shutil, tempfile, os


@router.get("/admin/import-schedule", response_class=HTMLResponse)
def import_schedule_page(request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    return templates.TemplateResponse("import_schedule.html", {"request": request, "result": None})


@router.post("/admin/import-schedule", response_class=HTMLResponse)
async def import_schedule(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    require_role(request, db, ROLE_ADMIN)
    contents = await file.read()

    # Сохраняем как schedule.xlsx
    with open(SCHEDULE_XLSX, "wb") as f:
        f.write(contents)

    # Загружаем в базу
    try:
        rows = _load_from_xlsx(SCHEDULE_XLSX)
        added = 0
        updated = 0
        from datetime import time as dtime
        import models as m
        for dir_name, slot_date, slot_time, capacity in rows:
            direction = db.query(m.Direction).filter_by(name=dir_name).first()
            if not direction:
                direction = m.Direction(name=dir_name)
                db.add(direction)
                db.flush()

            exists = db.query(m.Slot).filter_by(
                direction_id=direction.id,
                date=slot_date,
                time=slot_time,
            ).first()

            if not exists:
                db.add(m.Slot(
                    direction_id=direction.id,
                    date=slot_date,
                    time=slot_time,
                    capacity=capacity,
                ))
                added += 1
            else:
                exists.capacity = capacity
                updated += 1

        db.commit()
        result = {"added": added, "updated": updated, "error": None}
    except Exception as e:
        result = {"added": 0, "updated": 0, "error": str(e)}

    return templates.TemplateResponse("import_schedule.html", {
        "request": request,
        "result": result,
    })


@router.post("/admin/direction/{direction_id}/delete")
def delete_direction(direction_id: int, request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    d = db.query(models.Direction).filter_by(id=direction_id).first()
    if d:
        db.delete(d)
        db.commit()
    return RedirectResponse("/admin", status_code=302)


# ── Блокировка дней ───────────────────────────────────────────────────────────

@router.post("/admin/blocked-days/add")
def add_blocked_day(
    request: Request,
    block_date: str = Form(...),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    require_role(request, db, ROLE_ADMIN)
    try:
        target_date = datetime.strptime(block_date, "%Y-%m-%d").date()
        # Проверяем что день ещё не заблокирован
        exists = db.query(models.BlockedDay).filter_by(date=target_date).first()
        if not exists:
            db.add(models.BlockedDay(date=target_date, reason=reason.strip() or None))
            db.commit()
    except ValueError:
        pass
    return RedirectResponse("/admin", status_code=302)


@router.post("/admin/blocked-days/{day_id}/remove")
def remove_blocked_day(day_id: int, request: Request, db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    bd = db.query(models.BlockedDay).filter_by(id=day_id).first()
    if bd:
        db.delete(bd)
        db.commit()
    return RedirectResponse("/admin", status_code=302)


# ── Динамическая CSS тема (публичный эндпоинт — нужен всем страницам) ──────
@router.get("/api/theme.css")
def theme_css(db: Session = Depends(get_db)):
    theme_name = _get_setting(db, "theme", "indigo")
    p = THEMES.get(theme_name, THEMES["indigo"])
    css = f"""
:root {{
  --p50:  {p['50']};
  --p100: {p['100']};
  --p200: {p['200']};
  --p300: {p['300']};
  --p400: {p['400']};
  --p500: {p['500']};
  --p600: {p['600']};
  --p700: {p['700']};
  --p800: {p['800']};
  --p900: {p['900']};
}}
""".strip()
    return Response(content=css, media_type="text/css",
                    headers={"Cache-Control": "no-cache"})


# ── Сохранение темы (только для админа) ─────────────────────────────────────
@router.post("/admin/settings/theme")
def save_theme(request: Request, theme: str = Form(...), db: Session = Depends(get_db)):
    require_role(request, db, ROLE_ADMIN)
    if theme in THEMES:
        _set_setting(db, "theme", theme)
    return RedirectResponse("/admin?theme_saved=1", status_code=302)


# ---- Ручное добавление пользователя ----
@router.post("/admin/user/create")
def create_user(
    request: Request,
    username:  str = Form(...),
    password:  str = Form(...),
    full_name: str = Form(""),
    role:      str = Form("volunteer"),
    db: Session = Depends(get_db),
):
    require_role(request, db, ROLE_ADMIN)
    from services.auth import hash_password
    existing = db.query(models.User).filter_by(username=username.strip()).first()
    if existing:
        return RedirectResponse("/admin/import?error=exists", status_code=302)
    db.add(models.User(
        username=username.strip(),
        full_name=full_name.strip() or None,
        password_hash=hash_password(password),
        role=role,
        is_active=True,
    ))
    db.commit()
    return RedirectResponse("/admin/import?success=created", status_code=302)


# ── Отчёт для лотосов (экспорт посещаемости) ────────────────────────────────
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@router.get("/admin/lotos-report")
def lotos_report(request: Request, db: Session = Depends(get_db)):
    # Доступно админам и лотосам
    user = require_role(request, db, ROLE_ADMIN)  # проверит admin
    if not user:
        user = require_role(request, db, ROLE_LOTOS)

    # Запрос: все записи с посещаемостью
    rows = (
        db.query(
            models.Slot.date,
            models.Slot.time,
            models.Direction.name.label("direction"),
            models.User.full_name,
            models.User.username,
            models.Attendance.present,
        )
        .join(models.Booking, models.Slot.id == models.Booking.slot_id)
        .join(models.User, models.Booking.user_id == models.User.id)
        .join(models.Direction, models.Slot.direction_id == models.Direction.id)
        .outerjoin(models.Attendance, models.Attendance.booking_id == models.Booking.id)
        .order_by(models.Slot.date, models.Slot.time, models.Direction.name)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    headers = ["Дата", "Время", "Направление", "Волонтёр", "Логин", "Статус"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        status = "Пришёл" if row.present is True else ("Не пришёл" if row.present is False else "Не отмечен")
        ws.append([
            row.date.strftime("%d.%m.%Y"),
            row.time.strftime("%H:%M"),
            row.direction,
            row.full_name or "—",
            row.username,
            status,
        ])

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lotos_report.xlsx"},
    )


@router.get("/admin/login-logs", response_class=HTMLResponse)
def login_logs(
    request: Request,
    page: int = Query(default=1, ge=1),
    db: Session = Depends(get_db),
):
    """Страница логов входов пользователей."""
    require_role(request, db, ROLE_ADMIN)

    per_page = 50
    offset = (page - 1) * per_page

    # Получаем логи с информацией о пользователях
    logs = db.query(models.LoginLog).options(
        joinedload(models.LoginLog.user)
    ).order_by(models.LoginLog.created_at.desc()).offset(offset).limit(per_page).all()

    total = db.query(func.count(models.LoginLog.id)).scalar()
    total_pages = (total + per_page - 1) // per_page

    return templates.TemplateResponse("login_logs.html", {
        "request": request,
        "logs": logs,
        "page": page,
        "total_pages": total_pages,
        "total": total,
    })
