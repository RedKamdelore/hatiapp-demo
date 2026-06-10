import io
from datetime import datetime, time
from openpyxl import load_workbook
from database import SessionLocal
from config import BASE_DIR
import models

SCHEDULE_XLSX = BASE_DIR / "schedule.xlsx"


def _load_from_xlsx(path) -> list[tuple]:
    """Читает xlsx и возвращает список (direction_name, date, time, capacity)."""
    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        dir_name     = str(row[0]).strip()
        date_raw     = row[1]
        time_raw     = str(row[2]).strip() if row[2] else ""
        capacity_raw = row[3]

        # Дата — может быть объектом datetime или строкой
        if hasattr(date_raw, "date"):
            slot_date = date_raw.date() if hasattr(date_raw, "date") else date_raw
        elif isinstance(date_raw, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    slot_date = datetime.strptime(date_raw.strip(), fmt).date()
                    break
                except ValueError:
                    continue
            else:
                print(f"  ⚠ Строка {i}: не могу разобрать дату '{date_raw}', пропускаю")
                continue
        else:
            print(f"  ⚠ Строка {i}: пустая дата, пропускаю")
            continue

        # Время — строка "10:00" или объект time
        if hasattr(time_raw, "hour"):
            slot_time = time_raw
        else:
            try:
                h, m = map(int, str(time_raw).split(":"))
                slot_time = time(h, m)
            except Exception:
                print(f"  ⚠ Строка {i}: не могу разобрать время '{time_raw}', пропускаю")
                continue

        capacity = int(capacity_raw) if capacity_raw is not None else 0
        rows.append((dir_name, slot_date, slot_time, capacity))
    return rows


def seed_schedule():
    if not SCHEDULE_XLSX.exists():
        print(f"⚠ Файл {SCHEDULE_XLSX} не найден, расписание не загружено")
        return

    db = SessionLocal()
    try:
        rows = _load_from_xlsx(SCHEDULE_XLSX)
        added = 0
        for dir_name, slot_date, slot_time, capacity in rows:
            direction = db.query(models.Direction).filter_by(name=dir_name).first()
            if not direction:
                direction = models.Direction(name=dir_name)
                db.add(direction)
                db.flush()

            exists = db.query(models.Slot).filter_by(
                direction_id=direction.id,
                date=slot_date,
                time=slot_time,
            ).first()

            if not exists:
                db.add(models.Slot(
                    direction_id=direction.id,
                    date=slot_date,
                    time=slot_time,
                    capacity=capacity,
                ))
                added += 1
            else:
                # Обновляем вместимость если изменилась
                exists.capacity = capacity

        db.commit()
        print(f"✅ Расписание загружено ({added} новых слотов)")
    finally:
        db.close()
