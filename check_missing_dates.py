import openpyxl
from datetime import date, timedelta
import re

SCHEDULE_START = date(2026, 7, 9)
SCHEDULE_END = date(2026, 7, 13)

def parse_date_from_header(header):
    if not header or not isinstance(header, str):
        return None
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    match = re.search(r'(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)', header.lower())
    if match:
        day = int(match.group(1))
        month = months[match.group(2)]
        return date(2026, month, day)
    return None

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\Анкеты Хати 2026 (1).xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Карта колонок с датами
date_columns = {}
for h in headers:
    d = parse_date_from_header(h)
    if d:
        date_columns[h] = d

# ТГ которые есть в БД но без дат
missing_tgs = ['Pruzraki', 'NikiWay2', 'CookieZoya', 'Tweedn', 'Old_Monk_ey', 
               'DorianMatsui', 'CeleryBun', 'Polina_Belkin', 'rWbl_49', 
               'Io_Tkhorzh', 'Blg10001', 'Annetta_859', 'KaAnhlie', 'Dybrawka']
missing_tgs = [t.lower() for t in missing_tgs]

print(f"Проверка {len(missing_tgs)} волонтёров без дат:\n")

for row in ws.iter_rows(min_row=2, values_only=True):
    if len(row) > 5 and row[5]:
        tg = str(row[5]).strip().lower()
        if tg in missing_tgs:
            pozyvnoy = row[0] if len(row) > 0 else 'N/A'
            print(f"\n{tg} ({pozyvnoy}):")
            
            yes_dates = []
            for col_name, d in date_columns.items():
                col_idx = headers.index(col_name)
                if col_idx < len(row):
                    val = row[col_idx]
                    if val and str(val).strip().lower() == 'да':
                        yes_dates.append(d)
            
            if yes_dates:
                yes_dates.sort()
                print(f"  Даты с 'Да': {[str(d) for d in yes_dates]}")
                print(f"  Первая: {yes_dates[0]}, Последняя: {yes_dates[-1]}")
                print(f"  Первая >= SCHEDULE_END ({SCHEDULE_END})? {yes_dates[0] >= SCHEDULE_END}")
                print(f"  Последняя <= SCHEDULE_START ({SCHEDULE_START})? {yes_dates[-1] <= SCHEDULE_START}")
            else:
                print(f"  НЕТ ни одного 'Да' в датных колонках!")
                # Покажем что есть
                print(f"  Значения в датных колонках:")
                for col_name, d in sorted(date_columns.items(), key=lambda x: x[1]):
                    col_idx = headers.index(col_name)
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val:
                            print(f"    {d}: {val}")
