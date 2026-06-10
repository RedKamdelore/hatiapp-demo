#!/usr/bin/env python3
"""
Импорт волонтёров из Excel-анкет (финальная версия v2).
Проверяет дубликаты username перед созданием.
"""
import sys
sys.path.insert(0, r'C:\Users\Administrator\Desktop\САМОПАЛ\Hatiapp_cowork_OPENCODE')

import openpyxl
from datetime import date, timedelta
from database import SessionLocal
import models
import re

SCHEDULE_START = date(2026, 7, 9)
SCHEDULE_END = date(2026, 7, 13)

def parse_date_from_header(header):
    if not header or not isinstance(header, str):
        return None
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    match = re.search(r'(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)', header.lower())
    if match:
        day = int(match.group(1))
        month = months[match.group(2)]
        return date(2026, month, day)
    return None

def get_dates(row, date_columns, headers):
    yes_dates = []
    for col_name, d in date_columns.items():
        col_idx = headers.index(col_name)
        if col_idx < len(row):
            val = row[col_idx]
            if val and str(val).strip().lower() == 'да':
                yes_dates.append(d)
    if not yes_dates:
        return None, None
    yes_dates.sort()
    return yes_dates[0], yes_dates[-1] + timedelta(days=1)

def process():
    db = SessionLocal()
    wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\Анкеты Хати 2026 (1).xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    date_columns = {}
    for h in headers:
        d = parse_date_from_header(h)
        if d:
            date_columns[h] = d
    
    print(f"Найдено колонок с датами: {len(date_columns)}")
    
    # Получаем существующих пользователей
    existing_users = {u.username: u for u in db.query(models.User).all()}
    print(f"Существующих пользователей в БД: {len(existing_users)}")
    for u in existing_users.values():
        print(f"  {u.username} ({u.role})")
    
    created = 0
    skipped = 0
    blocked = []
    duplicates = []
    password_map = []
    
    from services.auth import hash_password
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            skipped += 1
            continue
            
        pozyvnoy = row[0]
        phone = row[4]
        tg = row[5]
        
        if not tg or not pozyvnoy:
            skipped += 1
            continue
        
        username = str(tg).strip()
        full_name = str(pozyvnoy).strip()
        password = str(phone).strip() if phone else 'changeme'
        
        # Проверяем дубликат username
        if username in existing_users:
            existing = existing_users[username]
            duplicates.append(f"{username}: уже есть как {existing.role} ({existing.full_name or 'N/A'})")
            continue
        
        arrival, departure = get_dates(row, date_columns, headers)
        
        if not arrival or not departure:
            blocked.append(f"{username} ({full_name}): нет дат")
            continue
        
        if arrival >= SCHEDULE_END or departure <= SCHEDULE_START:
            blocked.append(f"{username} ({full_name}): вне расписания ({arrival} - {departure})")
            continue
        
        # Создаём нового пользователя
        user = models.User(
            username=username,
            full_name=full_name,
            password_hash=hash_password(password),
            role='volunteer',
            is_active=True,
            arrival_date=arrival,
            departure_date=departure,
        )
        db.add(user)
        existing_users[username] = user  # Добавляем в кэш
        created += 1
        password_map.append((username, password, full_name, arrival, departure))
    
    db.commit()
    
    # Сохраняем пароли
    with open('volunteer_passwords.txt', 'w', encoding='utf-8') as f:
        f.write('username|password|full_name|arrival|departure\n')
        for item in password_map:
            f.write('|'.join(str(x) for x in item) + '\n')
    
    print("\n" + "="*60)
    print(f"Создано: {created}")
    print(f"Пропущено: {skipped}")
    print(f"Дубликатов (уже в БД): {len(duplicates)}")
    for d in duplicates:
        print(f"  • {d}")
    print(f"Заблокировано: {len(blocked)}")
    for b in blocked:
        print(f"  • {b}")
    print("="*60)
    
    # Проверим итог
    total = db.query(models.User).filter(models.User.role == 'volunteer').count()
    print(f"\nИтого волонтёров в БД: {total}")
    
    db.close()

if __name__ == "__main__":
    process()
