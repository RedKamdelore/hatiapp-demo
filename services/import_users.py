"""
Импорт волонтёров из Excel.
Ожидаемый формат таблицы:
  Колонка A: позывной (username)
  Колонка B: пароль
  Колонка C: имя (необязательно)
  Колонка D: роль (необязательно, по умолчанию volunteer)
  Колонка E: дата заезда (YYYY-MM-DD, необязательно)
  Колонка F: дата отъезда (YYYY-MM-DD, необязательно)
"""
import io
from openpyxl import load_workbook
from passlib.context import CryptContext
from sqlalchemy.orm import Session
import models

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")

from config import ROLE_ADMIN, ROLE_LEADER, ROLE_VOLUNTEER, ROLE_LOTOS, ROLE_PERMANENT

ALLOWED_ROLES = {ROLE_ADMIN, ROLE_LEADER, ROLE_VOLUNTEER, ROLE_LOTOS, ROLE_PERMANENT}


def import_users_from_excel(file_bytes: bytes, db: Session) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Пропускаем пустые строки
        if not row or not row[0]:
            continue

        username = str(row[0]).strip()
        password = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        full_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
        role      = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ROLE_VOLUNTEER

        if not username or not password:
            errors.append(f"Строка {i}: пустой логин или пароль")
            continue

        if role not in ALLOWED_ROLES:
            role = ROLE_VOLUNTEER

        # Парсим даты заезда/отъезда
        from datetime import datetime as _dt, date as _date
        arrival = None
        if len(row) > 4 and row[4]:
            try:
                if isinstance(row[4], _date):
                    arrival = row[4]
                elif isinstance(row[4], _dt):
                    arrival = row[4].date()
                else:
                    arrival = _dt.strptime(str(row[4]).strip(), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        departure = None
        if len(row) > 5 and row[5]:
            try:
                if isinstance(row[5], _date):
                    departure = row[5]
                elif isinstance(row[5], _dt):
                    departure = row[5].date()
                else:
                    departure = _dt.strptime(str(row[5]).strip(), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        exists = db.query(models.User).filter_by(username=username).first()
        if exists:
            skipped += 1
            continue

        db.add(models.User(
            username=username,
            full_name=full_name,
            password_hash=pwd_context.hash(password),
            role=role,
            is_active=True,
            arrival_date=arrival,
            departure_date=departure,
        ))
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
